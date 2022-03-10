from asyncio import SafeChildWatcher
import re
from importlib import resources
from turtle import down
from instagrapi import Client
from post import Post
from post import PostType
from openpyxl import formatting
from openpyxl import Workbook, styles
from openpyxl.drawing.image import Image
from openpyxl.utils import get_column_letter

def scrape(postAmount, account_name, account_password, scraping_account, saving_location, excel_sheet_output_name, starting_from_post_nr, download_pics):

    print(f"Post Amount: {postAmount}")
    print(f"Account Name: {account_name}")
    print(f"Account Password: {account_password}")
    print(f"Scraping Account: {scraping_account}")
    print(f"Saving Location: {saving_location}")
    print(f"Excel Sheet Output Name {excel_sheet_output_name}")

    #Setting Up Instagrapi Client
    cl = Client()
    cl.login(account_name, account_password)

    #Creating and Selecting Excel Workbook and Worksheet
    wb = Workbook()
    ws = wb.active

    #Loading Post-Data of Scraping Account (Usually kickz93)
    print("Getting Account...")
    scraping_account_user_id = cl.user_id_from_username(scraping_account)
    print("Getting Posts...")
    lastPost = int(postAmount) + int(starting_from_post_nr)
    scraping_account_posts = cl.user_medias(scraping_account_user_id, amount=lastPost)
    print("Finished getting posts...")
    del scraping_account_posts[:int(starting_from_post_nr)]
    #Empty List for Posts (Type: Post)
    post_list = []

    #Creating a new Post for every Element in post_data
    loadingProgress = 1

    for post_data in scraping_account_posts:
        newPost = Post(
            sc = post_data.code,
            dt = post_data.taken_at,
            l = post_data.like_count,
            com = post_data.comment_count,
            cap = post_data.caption_text,
            tn = getThumbnailString(post_data= post_data),
            p = int(post_data.pk),
            typ = PostType(post_data.media_type),
            ht = getHashtags(post_data.caption_text)
        )
        newPost.brand = newPost.findBrand(tags = newPost.hashtags)
        newPost.category = newPost.findCategory(tags=newPost.hashtags)

        if download_pics:
            downloadMedia(newPost, cl, saving_location)
        post_list.append(newPost)
        print(f"Post added...{loadingProgress}/{postAmount}")
        loadingProgress = loadingProgress + 1

    #Set-up and fill the Excel Sheet
    setUpRowTitles(ws)
    ws.column_dimensions[get_column_letter(1)].autosize = True
    column = 2
    for post in post_list:
        addPostToExcel(post = post, column = column, ws = ws, download_pics=download_pics)
        column = column + 1

    #Save Excel Sheet
    ws.row_dimensions[4].height = 100
    wb.save(filename= f"{saving_location}/{excel_sheet_output_name}.xlsx")
    print("Done!")



def getHashtags(caption:str):
    return re.findall(r"#(\w+)", caption)

def downloadMedia(post: Post, cl: Client, saving_location):
    if post.thumbnail == "Error":
        post.notDownloadable = True
    if (post.post_type == PostType.PHOTO) & (post.notDownloadable == False):
        post.photoLocation = cl.photo_download(media_pk= post.pk, folder = saving_location)
    else:
        if (post.post_type == PostType.ALBUM) & (post.notDownloadable == False):
            post.photoLocation = cl.photo_download_by_url(url = post.thumbnail, folder = saving_location)
        else:
            post.photoLocation = "images/Platzhalter.jpg"


def getThumbnailString(post_data):
    url = "Platzhalter.jpg"
    if post_data.media_type == 1:
        url = post_data.thumbnail_url
    else:
        if post_data.media_type == 8:
            for resource in post_data.resources:
                if resource.media_type == 1:
                    url = resource.thumbnail_url
                    break
                url = "Error"    
    return url

def addPostToExcel(post: Post, column: int, ws, download_pics: bool):
    if (post.post_type == PostType.PHOTO or post.post_type == PostType.ALBUM) & download_pics:
            image = Image(post.photoLocation)
            width = image.width
            factor = 100/width
            image.width = image.width * factor
            image.height = image.height * factor
            ws.add_image(img= image, anchor= f'{get_column_letter(column)}3')

    ws[f"{get_column_letter(column)}1"] = post.shortcode
    ws[f"{get_column_letter(column)}2"] = post.link
    ws[f"{get_column_letter(column)}3"] = post.datetime.isoformat()
    ws[f"{get_column_letter(column)}5"] = post.likes
    ws[f"{get_column_letter(column)}6"] = post.comments
    ws[f"{get_column_letter(column)}7"] = post.caption
    ws[f"{get_column_letter(column)}8"] = post.category
    ws[f"{get_column_letter(column)}9"] = post.brand
    row = 10
    for hashtag in post.hashtags:
        ws[f"{get_column_letter(column)}{row}"] = hashtag
        row = row + 1
    ws.column_dimensions[get_column_letter(column)].autosize = True


def setUpRowTitles(ws):
    ws["A1"] = "PostId"
    ws["A2"] = "Postlink"
    ws["A3"] = "Postdatum"
    ws["A4"] = "Title Image"
    ws["A5"] = "Likes"
    ws["A6"] = "Comments"
    ws["A7"] = "Caption"
    ws["A8"] = "Category"
    ws["A9"] = "Brand"
    ws["A10"] = "Hashtags"

if __name__ == "__main__":
    #Options
    starting_from_post_nr = 0
    postAmount = 20
    account_name = "instagrapi.finn2"
    account_password = "Instagrapi1993"
    scraping_account = "kickz93"
    saving_location = "/Users/finn/Kickz Werkstudent/Python Webscraper"
    excel_sheet_output_name = "Instagram Report März 22"
    download_pics = False
    
    scrape(
        starting_from_post_nr = starting_from_post_nr,
        postAmount=postAmount,
        account_name=account_name,
        account_password=account_password,
        scraping_account=scraping_account,
        saving_location=saving_location,
        excel_sheet_output_name=excel_sheet_output_name,
        download_pics= download_pics
    )
